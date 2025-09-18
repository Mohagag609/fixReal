"""
Report generation services for the real estate management system.
"""

import io
from datetime import datetime
from decimal import Decimal
from typing import Dict, Any, List, Optional
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment

from ..models import Customer, Unit, Contract, Installment
from .calculations import calculate_total_sales, calculate_unit_counts


class PDFReportGenerator:
    """PDF report generator."""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
    
    def generate_financial_report(self, start_date: datetime = None, end_date: datetime = None) -> HttpResponse:
        """Generate financial report in PDF format."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        # Get data
        sales_data = calculate_total_sales()
        
        # Build content
        story = []
        
        # Title
        title = Paragraph("التقرير المالي", self.styles['Heading1'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Financial summary table
        financial_data_table = [
            ['المؤشر', 'القيمة (ريال)'],
            ['إجمالي المبيعات', f"{sales_data['total_sales']:,.2f}"],
            ['إجمالي العقود', str(sales_data['total_contracts'])],
            ['العقود النشطة', str(sales_data['active_contracts'])],
        ]
        
        financial_table = Table(financial_data_table)
        financial_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(financial_table)
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="financial_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        return response


class ExcelReportGenerator:
    """Excel report generator."""
    
    def generate_financial_report(self, start_date: datetime = None, end_date: datetime = None) -> HttpResponse:
        """Generate financial report in Excel format."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "التقرير المالي"
        
        # Get data
        sales_data = calculate_total_sales()
        
        # Title
        ws['A1'] = "التقرير المالي"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Financial data
        ws['A3'] = "إجمالي المبيعات"
        ws['B3'] = f"{sales_data['total_sales']:,.2f}"
        ws['A4'] = "إجمالي العقود"
        ws['B4'] = str(sales_data['total_contracts'])
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="financial_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        wb.save(response)
        return response


def generate_report(report_type: str, format: str = 'pdf', **kwargs) -> HttpResponse:
    """Generate a report based on type and format."""
    if format.lower() == 'pdf':
        generator = PDFReportGenerator()
    elif format.lower() == 'excel':
        generator = ExcelReportGenerator()
    else:
        raise ValueError("Unsupported format. Use 'pdf' or 'excel'.")
    
    if report_type == 'financial':
        return generator.generate_financial_report(
            start_date=kwargs.get('start_date'),
            end_date=kwargs.get('end_date')
        )
    else:
        raise ValueError(f"Unsupported report type: {report_type}")