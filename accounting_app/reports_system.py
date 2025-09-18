"""
نظام التقارير المتقدم
Advanced Reports System

يحتوي على:
- منشئ التقارير الديناميكي
- تقارير مخصصة
- تصدير متعدد الصيغ (PDF, Excel, CSV, JSON)
- طباعة التقارير
- جدولة التقارير
- تقارير تفاعلية
"""

from django.db import models
from django.contrib.auth.models import User
from django.utils import timezone
from django.core.files.storage import default_storage
from django.conf import settings
from django.db.models import Q, Count, Sum, Avg, Max, Min
from datetime import datetime, timedelta
import json
import logging
from typing import List, Dict, Optional, Any
from enum import Enum
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import csv
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import base64

logger = logging.getLogger(__name__)


class ReportCategory(models.TextChoices):
    """فئات التقارير"""
    FINANCIAL = 'financial', 'مالية'
    CONTRACTS = 'contracts', 'عقود'
    CUSTOMERS = 'customers', 'عملاء'
    PARTNERS = 'partners', 'شركاء'
    UNITS = 'units', 'وحدات'
    PAYMENTS = 'payments', 'مدفوعات'
    SAFES = 'safes', 'خزنات'
    BROKERS = 'brokers', 'سماسرة'
    ANALYTICS = 'analytics', 'تحليلات'
    CUSTOM = 'custom', 'مخصص'


class ReportFormat(models.TextChoices):
    """صيغ التقارير"""
    PDF = 'pdf', 'PDF'
    EXCEL = 'excel', 'Excel'
    CSV = 'csv', 'CSV'
    JSON = 'json', 'JSON'
    HTML = 'html', 'HTML'


class ReportStatus(models.TextChoices):
    """حالات التقارير"""
    DRAFT = 'draft', 'مسودة'
    ACTIVE = 'active', 'نشط'
    SCHEDULED = 'scheduled', 'مجدول'
    RUNNING = 'running', 'قيد التشغيل'
    COMPLETED = 'completed', 'مكتمل'
    FAILED = 'failed', 'فشل'
    ARCHIVED = 'archived', 'مؤرشف'


class ReportTemplate(models.Model):
    """قوالب التقارير"""
    name = models.CharField(max_length=255, verbose_name="اسم القالب")
    description = models.TextField(blank=True, null=True, verbose_name="الوصف")
    category = models.CharField(max_length=50, choices=ReportCategory.choices, verbose_name="الفئة")
    query = models.TextField(verbose_name="استعلام قاعدة البيانات")
    fields = models.JSONField(default=list, verbose_name="الحقول")
    filters = models.JSONField(default=dict, verbose_name="الفلاتر")
    sorting = models.JSONField(default=dict, verbose_name="الترتيب")
    grouping = models.JSONField(default=dict, verbose_name="التجميع")
    calculations = models.JSONField(default=list, verbose_name="الحسابات")
    charts = models.JSONField(default=list, verbose_name="الرسوم البيانية")
    formatting = models.JSONField(default=dict, verbose_name="التنسيق")
    is_public = models.BooleanField(default=False, verbose_name="عام")
    is_active = models.BooleanField(default=True, verbose_name="نشط")
    created_by = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name="أنشأه")
    created_at = models.DateTimeField(auto_now_add=True, verbose_name="تاريخ الإنشاء")
    updated_at = models.DateTimeField(auto_now=True, verbose_name="تاريخ التحديث")

    class Meta:
        verbose_name = "قالب التقرير"
        verbose_name_plural = "قوالب التقارير"

    def __str__(self):
        return self.name


class ReportDefinition(models.Model):
    """تعريفات التقارير"""
    name = models.CharField(max_length=255, verbose_name="اسم التقرير")
    description = models.TextField(blank=True, null=True, verbose_name="الوصف")
    category = models.CharField(max_length=50, choices=ReportCategory.choices, verbose_name="الفئة")
    template = models.ForeignKey(ReportTemplate, on_delete=models.CASCADE, verbose_name="القالب")
    parameters = models.JSONField(default=dict, verbose_name="المعاملات")
    filters = models.JSONField(default=dict, verbose_name="الفلاتر")
    sorting = models.JSONField(default=dict, verbose_name="الترتيب")
    grouping = models.JSONField(default=dict, verbose_name="التجميع")
    calculations = models.JSONField(default=list, verbose_name="الحسابات")
    charts = models.JSONField(default=list, verbose_name="الرسوم البيانية")
    formatting = models.JSONField(default=dict, verbose_name="التنسيق")
    is_public = models.BooleanField(default=False, verbose_name="عام")
    is_active = models.BooleanField(default=True, verbose_name="نشط")
    created_by = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name="أنشأه")
    created_at = models.DateTimeField(auto_now_add=True, verbose_name="تاريخ الإنشاء")
    updated_at = models.DateTimeField(auto_now=True, verbose_name="تاريخ التحديث")

    class Meta:
        verbose_name = "تعريف التقرير"
        verbose_name_plural = "تعريفات التقارير"

    def __str__(self):
        return self.name


class ReportExecution(models.Model):
    """تنفيذ التقارير"""
    report_definition = models.ForeignKey(ReportDefinition, on_delete=models.CASCADE, verbose_name="تعريف التقرير")
    executed_by = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name="نفذه")
    parameters = models.JSONField(default=dict, verbose_name="المعاملات")
    filters = models.JSONField(default=dict, verbose_name="الفلاتر")
    status = models.CharField(max_length=20, choices=ReportStatus.choices, default=ReportStatus.DRAFT, verbose_name="الحالة")
    started_at = models.DateTimeField(null=True, blank=True, verbose_name="وقت البداية")
    completed_at = models.DateTimeField(null=True, blank=True, verbose_name="وقت الانتهاء")
    error_message = models.TextField(blank=True, null=True, verbose_name="رسالة الخطأ")
    file_path = models.CharField(max_length=500, blank=True, null=True, verbose_name="مسار الملف")
    file_size = models.PositiveIntegerField(default=0, verbose_name="حجم الملف")
    row_count = models.PositiveIntegerField(default=0, verbose_name="عدد الصفوف")
    execution_time = models.FloatField(default=0, verbose_name="وقت التنفيذ")
    metadata = models.JSONField(default=dict, verbose_name="البيانات الإضافية")
    created_at = models.DateTimeField(auto_now_add=True, verbose_name="تاريخ الإنشاء")

    class Meta:
        verbose_name = "تنفيذ التقرير"
        verbose_name_plural = "تنفيذ التقارير"
        ordering = ['-created_at']

    def __str__(self):
        return f"{self.report_definition.name} - {self.get_status_display()}"


class ReportSchedule(models.Model):
    """جدولة التقارير"""
    name = models.CharField(max_length=255, verbose_name="اسم الجدولة")
    report_definition = models.ForeignKey(ReportDefinition, on_delete=models.CASCADE, verbose_name="تعريف التقرير")
    schedule_type = models.CharField(max_length=20, choices=[
        ('daily', 'يومي'),
        ('weekly', 'أسبوعي'),
        ('monthly', 'شهري'),
        ('quarterly', 'ربعي'),
        ('yearly', 'سنوي'),
        ('custom', 'مخصص'),
    ], verbose_name="نوع الجدولة")
    schedule_config = models.JSONField(default=dict, verbose_name="إعدادات الجدولة")
    recipients = models.JSONField(default=list, verbose_name="المستقبلون")
    formats = models.JSONField(default=list, verbose_name="الصيغ")
    is_active = models.BooleanField(default=True, verbose_name="نشط")
    last_run = models.DateTimeField(null=True, blank=True, verbose_name="آخر تشغيل")
    next_run = models.DateTimeField(null=True, blank=True, verbose_name="التشغيل التالي")
    created_by = models.ForeignKey(User, on_delete=models.CASCADE, verbose_name="أنشأه")
    created_at = models.DateTimeField(auto_now_add=True, verbose_name="تاريخ الإنشاء")

    class Meta:
        verbose_name = "جدولة التقرير"
        verbose_name_plural = "جدولة التقارير"

    def __str__(self):
        return f"{self.name} - {self.get_schedule_type_display()}"


class ReportBuilder:
    """منشئ التقارير"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def create_report(self, 
                     name: str,
                     description: str,
                     category: str,
                     query: str,
                     fields: List[Dict],
                     filters: Dict = None,
                     sorting: Dict = None,
                     grouping: Dict = None,
                     calculations: List[Dict] = None,
                     charts: List[Dict] = None,
                     formatting: Dict = None,
                     created_by: User = None) -> ReportTemplate:
        """إنشاء قالب تقرير جديد"""
        
        if filters is None:
            filters = {}
        
        if sorting is None:
            sorting = {}
        
        if grouping is None:
            grouping = {}
        
        if calculations is None:
            calculations = []
        
        if charts is None:
            charts = []
        
        if formatting is None:
            formatting = {}
        
        return ReportTemplate.objects.create(
            name=name,
            description=description,
            category=category,
            query=query,
            fields=fields,
            filters=filters,
            sorting=sorting,
            grouping=grouping,
            calculations=calculations,
            charts=charts,
            formatting=formatting,
            created_by=created_by
        )
    
    def execute_report(self, 
                      report_definition: ReportDefinition,
                      parameters: Dict = None,
                      filters: Dict = None,
                      executed_by: User = None) -> ReportExecution:
        """تنفيذ تقرير"""
        
        if parameters is None:
            parameters = {}
        
        if filters is None:
            filters = {}
        
        # إنشاء سجل التنفيذ
        execution = ReportExecution.objects.create(
            report_definition=report_definition,
            executed_by=executed_by,
            parameters=parameters,
            filters=filters,
            status=ReportStatus.RUNNING,
            started_at=timezone.now()
        )
        
        try:
            # تنفيذ الاستعلام
            data = self._execute_query(report_definition, parameters, filters)
            
            # تطبيق الحسابات
            if report_definition.calculations:
                data = self._apply_calculations(data, report_definition.calculations)
            
            # تطبيق التجميع
            if report_definition.grouping:
                data = self._apply_grouping(data, report_definition.grouping)
            
            # تطبيق الترتيب
            if report_definition.sorting:
                data = self._apply_sorting(data, report_definition.sorting)
            
            # حفظ النتائج
            execution.row_count = len(data)
            execution.completed_at = timezone.now()
            execution.status = ReportStatus.COMPLETED
            execution.execution_time = (execution.completed_at - execution.started_at).total_seconds()
            execution.save()
            
            return execution
            
        except Exception as e:
            self.logger.error(f"Error executing report {report_definition.id}: {str(e)}")
            execution.status = ReportStatus.FAILED
            execution.error_message = str(e)
            execution.completed_at = timezone.now()
            execution.save()
            
            return execution
    
    def _execute_query(self, report_definition: ReportDefinition, parameters: Dict, filters: Dict) -> List[Dict]:
        """تنفيذ استعلام قاعدة البيانات"""
        # هذا مثال مبسط - في التطبيق الحقيقي ستحتاج لمعالج استعلامات أكثر تعقيداً
        from django.db import connection
        
        query = report_definition.template.query
        
        # تطبيق المعاملات
        for key, value in parameters.items():
            query = query.replace(f"{{{key}}}", str(value))
        
        # تطبيق الفلاتر
        if filters:
            where_conditions = []
            for field, value in filters.items():
                if value:
                    where_conditions.append(f"{field} = '{value}'")
            
            if where_conditions:
                if "WHERE" in query.upper():
                    query += " AND " + " AND ".join(where_conditions)
                else:
                    query += " WHERE " + " AND ".join(where_conditions)
        
        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            results = cursor.fetchall()
            
            return [dict(zip(columns, row)) for row in results]
    
    def _apply_calculations(self, data: List[Dict], calculations: List[Dict]) -> List[Dict]:
        """تطبيق الحسابات على البيانات"""
        for calc in calculations:
            calc_type = calc.get('type')
            field = calc.get('field')
            formula = calc.get('formula')
            
            if calc_type == 'sum':
                total = sum(row.get(field, 0) for row in data if isinstance(row.get(field), (int, float)))
                for row in data:
                    row[f'{field}_total'] = total
            
            elif calc_type == 'average':
                values = [row.get(field, 0) for row in data if isinstance(row.get(field), (int, float))]
                avg = sum(values) / len(values) if values else 0
                for row in data:
                    row[f'{field}_average'] = avg
            
            elif calc_type == 'percentage':
                total = sum(row.get(field, 0) for row in data if isinstance(row.get(field), (int, float)))
                for row in data:
                    value = row.get(field, 0)
                    row[f'{field}_percentage'] = (value / total * 100) if total > 0 else 0
        
        return data
    
    def _apply_grouping(self, data: List[Dict], grouping: Dict) -> List[Dict]:
        """تطبيق التجميع على البيانات"""
        group_by = grouping.get('fields', [])
        if not group_by:
            return data
        
        # تجميع البيانات
        grouped = {}
        for row in data:
            key = tuple(row.get(field, '') for field in group_by)
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(row)
        
        # إنشاء البيانات المجمعة
        result = []
        for key, group in grouped.items():
            summary = {}
            for field in group_by:
                summary[field] = key[group_by.index(field)]
            
            # حساب الإجماليات
            for field in grouping.get('aggregations', []):
                field_name = field.get('field')
                agg_type = field.get('type')
                
                if agg_type == 'sum':
                    summary[f'{field_name}_sum'] = sum(row.get(field_name, 0) for row in group if isinstance(row.get(field_name), (int, float)))
                elif agg_type == 'count':
                    summary[f'{field_name}_count'] = len(group)
                elif agg_type == 'average':
                    values = [row.get(field_name, 0) for row in group if isinstance(row.get(field_name), (int, float))]
                    summary[f'{field_name}_avg'] = sum(values) / len(values) if values else 0
            
            result.append(summary)
        
        return result
    
    def _apply_sorting(self, data: List[Dict], sorting: Dict) -> List[Dict]:
        """تطبيق الترتيب على البيانات"""
        sort_fields = sorting.get('fields', [])
        if not sort_fields:
            return data
        
        # ترتيب البيانات
        for field in reversed(sort_fields):
            field_name = field.get('field')
            direction = field.get('direction', 'asc')
            
            data.sort(key=lambda x: x.get(field_name, ''), reverse=(direction == 'desc'))
        
        return data
    
    def export_to_pdf(self, data: List[Dict], report_definition: ReportDefinition, filename: str = None) -> str:
        """تصدير التقرير إلى PDF"""
        if filename is None:
            filename = f"report_{report_definition.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        file_path = f"reports/pdf/{filename}"
        full_path = default_storage.path(file_path)
        
        # إنشاء PDF
        doc = SimpleDocTemplate(full_path, pagesize=A4)
        elements = []
        
        # العنوان
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=getSampleStyleSheet()['Title'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        elements.append(Paragraph(report_definition.name, title_style))
        elements.append(Spacer(1, 12))
        
        # البيانات
        if data:
            # رؤوس الأعمدة
            headers = list(data[0].keys())
            
            # تحضير البيانات
            table_data = [headers]
            for row in data:
                table_data.append([str(row.get(header, '')) for header in headers])
            
            # إنشاء الجدول
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        
        # بناء PDF
        doc.build(elements)
        
        return file_path
    
    def export_to_excel(self, data: List[Dict], report_definition: ReportDefinition, filename: str = None) -> str:
        """تصدير التقرير إلى Excel"""
        if filename is None:
            filename = f"report_{report_definition.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        file_path = f"reports/excel/{filename}"
        full_path = default_storage.path(file_path)
        
        # إنشاء Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        if data:
            # رؤوس الأعمدة
            headers = list(data[0].keys())
            
            # كتابة الرؤوس
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # كتابة البيانات
            for row_idx, row in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(header, ''))
            
            # تنسيق الأعمدة
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(full_path)
        
        return file_path
    
    def export_to_csv(self, data: List[Dict], report_definition: ReportDefinition, filename: str = None) -> str:
        """تصدير التقرير إلى CSV"""
        if filename is None:
            filename = f"report_{report_definition.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        file_path = f"reports/csv/{filename}"
        full_path = default_storage.path(file_path)
        
        if data:
            with open(full_path, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = list(data[0].keys())
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)
        
        return file_path
    
    def export_to_json(self, data: List[Dict], report_definition: ReportDefinition, filename: str = None) -> str:
        """تصدير التقرير إلى JSON"""
        if filename is None:
            filename = f"report_{report_definition.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        file_path = f"reports/json/{filename}"
        full_path = default_storage.path(file_path)
        
        with open(full_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(data, jsonfile, ensure_ascii=False, indent=2)
        
        return file_path


# وظائف مساعدة للتقارير المالية
def create_financial_report_template() -> ReportTemplate:
    """إنشاء قالب تقرير مالي"""
    builder = ReportBuilder()
    
    return builder.create_report(
        name="التقرير المالي الشامل",
        description="تقرير مالي شامل لجميع المعاملات والأرصدة",
        category=ReportCategory.FINANCIAL,
        query="""
            SELECT 
                s.name as safe_name,
                s.balance as current_balance,
                COUNT(v.id) as transaction_count,
                SUM(CASE WHEN v.type = 'receipt' THEN v.amount ELSE 0 END) as total_receipts,
                SUM(CASE WHEN v.type = 'payment' THEN v.amount ELSE 0 END) as total_payments
            FROM accounting_app_safe s
            LEFT JOIN accounting_app_voucher v ON s.id = v.safe_id
            WHERE s.deleted_at IS NULL
            GROUP BY s.id, s.name, s.balance
        """,
        fields=[
            {"name": "safe_name", "label": "اسم الخزنة", "type": "text"},
            {"name": "current_balance", "label": "الرصيد الحالي", "type": "currency"},
            {"name": "transaction_count", "label": "عدد المعاملات", "type": "number"},
            {"name": "total_receipts", "label": "إجمالي الإيرادات", "type": "currency"},
            {"name": "total_payments", "label": "إجمالي المدفوعات", "type": "currency"},
        ],
        calculations=[
            {"type": "sum", "field": "current_balance"},
            {"type": "sum", "field": "total_receipts"},
            {"type": "sum", "field": "total_payments"},
        ]
    )


def create_contracts_report_template() -> ReportTemplate:
    """إنشاء قالب تقرير العقود"""
    builder = ReportBuilder()
    
    return builder.create_report(
        name="تقرير العقود",
        description="تقرير شامل لجميع العقود والأقساط",
        category=ReportCategory.CONTRACTS,
        query="""
            SELECT 
                c.id as contract_id,
                u.code as unit_code,
                cu.name as customer_name,
                c.total_price,
                c.down_payment,
                c.discount_amount,
                c.start as contract_date,
                COUNT(i.id) as installment_count,
                SUM(CASE WHEN i.status = 'مدفوع' THEN i.amount ELSE 0 END) as paid_amount,
                SUM(CASE WHEN i.status = 'غير مدفوع' THEN i.amount ELSE 0 END) as pending_amount
            FROM accounting_app_contract c
            JOIN accounting_app_unit u ON c.unit_id = u.id
            JOIN accounting_app_customer cu ON c.customer_id = cu.id
            LEFT JOIN accounting_app_installment i ON u.id = i.unit_id
            WHERE c.deleted_at IS NULL
            GROUP BY c.id, u.code, cu.name, c.total_price, c.down_payment, c.discount_amount, c.start
        """,
        fields=[
            {"name": "contract_id", "label": "رقم العقد", "type": "text"},
            {"name": "unit_code", "label": "كود الوحدة", "type": "text"},
            {"name": "customer_name", "label": "اسم العميل", "type": "text"},
            {"name": "total_price", "label": "السعر الإجمالي", "type": "currency"},
            {"name": "down_payment", "label": "المقدم", "type": "currency"},
            {"name": "discount_amount", "label": "الخصم", "type": "currency"},
            {"name": "contract_date", "label": "تاريخ العقد", "type": "date"},
            {"name": "installment_count", "label": "عدد الأقساط", "type": "number"},
            {"name": "paid_amount", "label": "المدفوع", "type": "currency"},
            {"name": "pending_amount", "label": "المتبقي", "type": "currency"},
        ],
        calculations=[
            {"type": "sum", "field": "total_price"},
            {"type": "sum", "field": "paid_amount"},
            {"type": "sum", "field": "pending_amount"},
        ]
    )


def create_partners_report_template() -> ReportTemplate:
    """إنشاء قالب تقرير الشركاء"""
    builder = ReportBuilder()
    
    return builder.create_report(
        name="تقرير الشركاء",
        description="تقرير شامل لجميع الشركاء وحصصهم",
        category=ReportCategory.PARTNERS,
        query="""
            SELECT 
                p.name as partner_name,
                p.phone,
                COUNT(up.id) as unit_count,
                SUM(up.percentage) as total_percentage,
                SUM(u.total_price * up.percentage / 100) as total_investment
            FROM accounting_app_partner p
            LEFT JOIN accounting_app_unitpartner up ON p.id = up.partner_id
            LEFT JOIN accounting_app_unit u ON up.unit_id = u.id
            WHERE p.deleted_at IS NULL
            GROUP BY p.id, p.name, p.phone
        """,
        fields=[
            {"name": "partner_name", "label": "اسم الشريك", "type": "text"},
            {"name": "phone", "label": "الهاتف", "type": "text"},
            {"name": "unit_count", "label": "عدد الوحدات", "type": "number"},
            {"name": "total_percentage", "label": "إجمالي النسبة", "type": "percentage"},
            {"name": "total_investment", "label": "إجمالي الاستثمار", "type": "currency"},
        ],
        calculations=[
            {"type": "sum", "field": "total_investment"},
            {"type": "average", "field": "total_percentage"},
        ]
    )